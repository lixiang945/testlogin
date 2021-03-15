import json

from config import *
import os
import openpyxl


class ReadEX:
    def __init__(self, filename, sheetname):
        self.file = base_path + os.sep + 'data' + os.sep + filename
        self.op = openpyxl.load_workbook(self.file)

        self.sh = self.op[sheetname]

        self.rows = self.sh.max_row
        self.col = self.sh.max_column
        print(f'rows is {self.rows}, col is {self.col}')

    # 2、读取excel
    def read_excel(self,casename):
        # 1、新建空列表（存储每行数据）
        case = list()
        all_case = list()
        # all_datas = dict()
        # datas =dict()
        for x in range(2, self.rows+1):
            data = dict()
            if str(self.sh.cell(x, excel_cell.get('casename')).value) == 'None':
                continue
            else:
                title = str(self.sh.cell(x, excel_cell.get('casename')).value)
            print(title)
            # for i in range(x, self.rows+1):
            #     data = dict()
            #     data['step'] = str(self.sh.cell(i, excel_cell.get('step')).value)
            #     data['method'] = str(self.sh.cell(i, excel_cell.get('method')).value)
            #     data['location'] = str(self.sh.cell(i, excel_cell.get('location')).value)
            #     case.append(data)
            #     end = str(self.sh.cell(i, excel_cell.get('end')).value)
            #     if end == 'None':
            #         pass
            #     else:
            #         datas = {
            #             'title': title,
            #             'cases': case
            #         }
            #         all_case.append(datas)
            #         print(all_case)
            #         # self.white_json(datas, f'{casename}.json')
            #         case.clear()
            #         break



 # 3、写入excel:
    def white_excel(self, x_y, msg):
        try:
            # x_y为一个列表，如[2,3]
            self.sh.cell(x_y[0], x_y[1]).value = msg
        except Exception as e:
            self.sh.cell(x_y[0], x_y[1]).value = e
        finally:
            self.op.save(self.file)
            self.op.close()

    # 4、读取Json
    def read_json(self, filename):
        file_name = base_path + os.sep + "data" + os.sep + filename + '.json'
        with open(file_name, 'r', encoding='utf-8') as f:
            # json.loads是将字符串转换为字典，load是读取内容
            return json.load(f)

    # 5、写入Json
    def white_json(self, case, filename):
        file_name = base_path + os.sep + "data" + os.sep + filename
        with open(file_name, 'w', encoding='utf-8') as f:
            # dump为写，dumps为转换
            json.dump(case, f, indent=4, ensure_ascii=False)


if __name__ == '__main__':
    rs = ReadEX('测试用例.xlsx', 'login1')
    rs.read_excel('123')

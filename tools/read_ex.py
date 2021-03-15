import json
# from tools.keywords import Keywords
from config import *
import os
import openpyxl
import time


class ReadEX:
    def __init__(self, filename, sheetname):
        self.file = base_path + os.sep + 'data' + os.sep + filename
        self.op = openpyxl.load_workbook(self.file)

        self.sh = self.op[sheetname]

        self.rows = self.sh.max_row
        self.col = self.sh.max_column
        # print(self.rows, self.col)

    # 2、读取excel
    def read_excel(self, casename):
        # 定义模块
        modelDict = dict()
        # 定义功能名feature name
        modelName = self.sh.cell(2, 1).value
        # 定义最外层列表，也就是modelName对应的值
        checkCase = list()
        modelDict[modelName] = checkCase
        # 定义cases外层的列表
        cList = list()
        # 定义一个列表，检查要创建的title是否存在
        for i in range(2, self.rows + 1):
            try:
                caseName = self.sh.cell(i, 2).value
                if caseName is not None and caseName not in checkCase:
                    # 定义包裹title和cases外层的字典
                    cnDict = dict()
                    checkCase.append(cnDict)
                    if caseName not in cnDict.keys():
                        cList = list()
                        cnDict['title'] = caseName
                        cnDict['cases'] = cList
                rowsCase = dict()
                cList.append(rowsCase)
                rowsCase['step'] = str(self.sh.cell(i, excel_cell.get('step')).value)
                rowsCase['x_y'] = [i, excel_cell.get("result")]
                rowsCase['desc'] = [i, excel_cell.get("desc")]
                rowsCase['method'] = str(self.sh.cell(i, excel_cell.get('method')).value)
                rowsCase['location'] = str(self.sh.cell(i, excel_cell.get('location')).value)
                params = str(self.sh.cell(i, excel_cell.get('params')).value)
                if params == 'None':
                    continue
                else:
                    rowsCase['params'] = str(self.sh.cell(i, excel_cell.get('params')).value)
            except Exception as e:
                self.white_excel([i, excel_cell.get("desc")], e)
        # print(modelDict)
        self.white_json(modelDict, f"{casename}.json")

    # 3、写入excel:
    def white_excel(self, x_y, msg):
        try:
            # x_y为一个列表，如[2,3]
            self.sh.cell(x_y[0], x_y[1]).value = msg
        except Exception as e:
            self.sh.cell(x_y[0], x_y[1]).value = e
        finally:
            self.op.save(self.file)
            self.op.close()

    # 4、读取Json
    def read_json(self, filename):
        file_name = base_path + os.sep + "data" + os.sep + filename + '.json'
        with open(file_name, 'r', encoding='utf-8') as f:
            # json.loads是将字符串转换为字典，load是读取内容
            return json.load(f)

    # 5、写入Json
    def white_json(self, case, filename):
        file_name = base_path + os.sep + "data" + os.sep + filename
        with open(file_name, 'w', encoding='utf-8') as f:
            # dump为写，dumps为转换
            json.dump(case, f, indent=4, ensure_ascii=False)


if __name__ == '__main__':
    rt = ReadEX('测试用例.xlsx', 'login1')
    print(rt.read_excel('test1'))

import json
# from tools.keywords import Keywords
from config import *
import os
import openpyxl
import time


class ReadEX:
    def __init__(self, filename, sheetname):
        self.file = base_path + os.sep + 'data' + os.sep + filename
        self.op = openpyxl.load_workbook(self.file)

        self.sh = self.op[sheetname]

        self.rows = self.sh.max_row
        self.col = self.sh.max_column
        # print(self.rows, self.col)

    # 2、读取excel
    def read_excel(self, casename):
        # 1、新建空列表（存储每行数据）
        case = list()
        all_case = list()
        datas = dict()
        all_datas = dict()
        for i in range(2, self.rows + 1):
            # 新建空字典
            data = dict()
            # 判断是否执行
            try:
                #  读取数据，追加到字典
                data['x_y'] = [i, excel_cell.get("result")]
                # 将异常写入备注
                data['desc'] = [i, excel_cell.get("desc")]
                data['assert'] = str(self.sh.cell(i, excel_cell.get('assert')).value)
                data['step'] = str(self.sh.cell(i, excel_cell.get('step')).value)
                data['method'] = str(self.sh.cell(i, excel_cell.get('method')).value)
                data['location'] = str(self.sh.cell(i, excel_cell.get('location')).value)
                if str(self.sh.cell(i, excel_cell.get('params')).value) == 'None':
                    pass
                else:
                    data['params'] = str(self.sh.cell(i, excel_cell.get('params')).value)
                # 将字段追加到空列表
                case.append(data)
                if str(self.sh.cell(2, excel_cell.get('casename')).value) == 'None':
                    pass
                else:
                    datas = {
                        "title": str(self.sh.cell(2, excel_cell.get('casename')).value),
                        "cases": case
                    }

                # 将读取结果写入excel
                # self.white_excel([i, cell_config.get("desc")], "data read is ok!")

            except Exception as e:
                self.white_excel([i, excel_cell.get("desc")], e)
            # 3、将列表数据写入json

        all_case.append(datas)
        all_datas = {
            str(self.sh.cell(2, excel_cell.get('feature')).value): all_case
        }
        self.white_json(all_datas, f"{casename}.json")
        return all_datas

    # 3、写入excel:
    def white_excel(self, x_y, msg):
        try:
            # x_y为一个列表，如[2,3]
            self.sh.cell(x_y[0], x_y[1]).value = msg
        except Exception as e:
            self.sh.cell(x_y[0], x_y[1]).value = e
        finally:
            self.op.save(self.file)
            self.op.close()

    # 4、读取Json
    def read_json(self, filename):
        file_name = base_path + os.sep + "data" + os.sep + filename + '.json'
        with open(file_name, 'r', encoding='utf-8') as f:
            # json.loads是将字符串转换为字典，load是读取内容
            return json.load(f)

    # 5、写入Json
    def white_json(self, case, filename):
        file_name = base_path + os.sep + "data" + os.sep + filename
        with open(file_name, 'w', encoding='utf-8') as f:
            # dump为写，dumps为转换
            json.dump(case, f, indent=4, ensure_ascii=False)

    def test_read(self, casename):
        # 定义模块
        modelDict = dict()
        # 定义功能名feature name
        modelName = self.sh.cell(2, 1).value
        # 定义最外层列表，也就是modelName对应的值
        checkCase = list()
        modelDict[modelName] = checkCase
        # 定义cases外层的列表
        cList = list()
        # 定义一个列表，检查要创建的title是否存在
        for i in range(2, self.rows+1):
            try:
                caseName = self.sh.cell(i, 2).value
                if caseName is not None and caseName not in checkCase:
                    # 定义包裹title和cases外层的字典
                    cnDict = dict()
                    checkCase.append(cnDict)
                    if caseName not in cnDict.keys():
                        cList = list()
                        cnDict['title'] = caseName
                        cnDict['cases'] = cList
                rowsCase = dict()
                cList.append(rowsCase)
                rowsCase['step'] = str(self.sh.cell(i, excel_cell.get('step')).value)
                rowsCase['x_y'] = [i, excel_cell.get("result")]
                rowsCase['desc'] = [i, excel_cell.get("desc")]
                rowsCase['method'] = str(self.sh.cell(i, excel_cell.get('method')).value)
                rowsCase['location'] = str(self.sh.cell(i, excel_cell.get('location')).value)
                params = str(self.sh.cell(i, excel_cell.get('params')).value)
                if params == 'None':
                    continue
                else:
                    rowsCase['params'] = str(self.sh.cell(i, excel_cell.get('params')).value)
            except Exception as e:
                self.white_excel([i, excel_cell.get("desc")], e)
        # print(modelDict)
        self.white_json(modelDict, f"{casename}.json")


if __name__ == '__main__':
    rt = ReadEX('测试用例.xlsx', 'logintest')
    # print(rt.read_excel('test1'))
    rt.test_read('123')
    # rt.read_excel('test1')
    # rt.read_result('result')

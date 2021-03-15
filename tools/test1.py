import json
import config
from config import *
import os
import openpyxl


class ReadEX:
    def __init__(self, filename, sheetname):
        self.file = base_path + os.sep + 'data' + os.sep + filename
        self.op = openpyxl.load_workbook(self.file)

        self.sh = self.op[sheetname]

        self.rows = self.sh.max_row
        self.col = self.sh.max_column
        print(f'rows is {self.rows}, col is {self.col}')

    # 2、读取excel
    def read_excel(self, casename):
        # 模块名字典
        modelDict = dict()
        # 用例数组,用来检查用例名是否存在
        caseList = list()
        # 模块名
        modelName = self.sh.cell(2, 1).value
        # 用例数组加入模块字典
        modelDict[modelName] = caseList
        cList = list()
        # 1.行遍历，x:行号
        for x in range(2, self.rows + 1):
            # 2.用例名设置
            caseName = self.sh.cell(x, 2).value
            if caseName not in caseList:
                if caseName is not None:
                    caseDict = dict()
                    caseList.append(caseDict)
                    if caseName not in caseDict.keys():
                        # 列集合
                        cList = list()
                        caseDict['title'] = caseName
                        caseDict['cases'] = cList
            # 列字典
            cDict = dict()
            cList.append(cDict)
            cDict['x_y'] = [x, excel_cell.get("result")]
            # 将异常写入备注
            cDict['desc'] = [x, excel_cell.get("desc")]
            cDict['step'] = str(self.sh.cell(x, excel_cell.get('step')).value)
            cDict['method'] = str(self.sh.cell(x, excel_cell.get('method')).value)
            cDict['location'] = str(self.sh.cell(x, excel_cell.get('location')).value)
            params = str(self.sh.cell(x, excel_cell.get('params')).value)
            if params == 'None':
                continue
            else:
                cDict['params'] = str(self.sh.cell(x, excel_cell.get('params')).value)


        # 4. 输出结果
        # print(modelDict)
        self.white_json(modelDict, f"{casename}.json")

    # 3、写入excel:
    def white_excel(self, x_y, msg):
        try:
            # x_y为一个列表，如[2,3]
            self.sh.cell(x_y[0], x_y[1]).value = msg
        except Exception as e:
            self.sh.cell(x_y[0], x_y[1]).value = e
        finally:
            self.op.save(self.file)
            self.op.close()

    # 4、读取Json
    def read_json(self, filename):
        file_name = base_path + os.sep + "data" + os.sep + filename + '.json'
        with open(file_name, 'r', encoding='utf-8') as f:
            # json.loads是将字符串转换为字典，load是读取内容
            return json.load(f)

    # 5、写入Json
    def white_json(self, case, filename):
        file_name = base_path + os.sep + "data" + os.sep + filename
        with open(file_name, 'w', encoding='utf-8') as f:
            # dump为写，dumps为转换
            json.dump(case, f, indent=4, ensure_ascii=False)


if __name__ == '__main__':
    rs = ReadEX('测试用例.xlsx', 'logintest')
    print(rs.read_excel('321'))
